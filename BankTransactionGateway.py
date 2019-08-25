import openpyxl
from openpyxl import load_workbook
wb=load_workbook('SITIP19_KIIT_Python_FinalProject1_Workbook.xlsx')
sheet=wb['UserDetails']
for i in range(1,8):
    sheet.cell(row=1,column=i).value=input("enter value in column {}".format(i))
import random,time
import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib
try:
    now = datetime.datetime.now()
    chance=3
    chance1=3
    chance2=3
    select1=input("Please Enter your Full Name or user ID")
    if select1.isnumeric():
        userid=select1
        for i in range(2,7):
            if sheet.cell(row=i,column=2).value==int(userid):
                flag=1
                position=i
                break
            else:
                flag=0
    else:
        name=select1.title()
        for i in range(2,7):
            if sheet.cell(row=i,column=4).value==str(name):
                flag=1
                position=i
                break
            else:
                flag=0
    if flag==0:
        print("Invalid User")
    else:
        if sheet.cell(row=position,column=7).value=='Blocked':
            mailID=sheet.cell(row=position,column=5).value
            op=input("Sorry! but your account was blocked.To reactivate press 1")
            if op=="1":
                otp=random.randint(1000,9999)
                print("Please wait..")
                fromaddr=""                            #Enter your email inside quotes
                toaddr=mailID
                msg=MIMEMultipart()
                msg['From']=fromaddr
                msg['To']=toaddr
                msg['Subject']='Account Verification'
                body="Sir,this is your OTP Number: {}".format(otp)
                msg.attach(MIMEText(body,'plain'))
                server=smtplib.SMTP('smtp.gmail.com',port=587)
                server.starttls()
                server.login(fromaddr,'')             #enter your email's password inside quotes
                text=msg.as_string()
                server.sendmail(fromaddr,toaddr,text)
                server.quit()
                time.sleep(5)
                print("An OTP has been sent to your mail. Plese Check!")
                while True:
                    otpv=int(input("Sir please enter the OTP that you received via mail: "))
                    if otpv==otp:
                        sheet.cell(position,7).value='Active'
                        wb.save('SITIP19_KIIT_Python_FinalProject1_Workbook.xlsx')
                        print("Valid OTP! Account activated")
                        flag1=1
                        break
                    else:
                        print("Invalid OTP")
        else:
            flag1=1
        if flag1==1:
            while chance!=0:
                accno=input("Enter account number: ")
                if sheet.cell(row=position,column=3).value==accno:
                    userID=sheet.cell(row=position,column=2).value
                    AccNo=sheet.cell(row=position,column=3).value
                    fullname=sheet.cell(row=position,column=4).value
                    mailID=sheet.cell(row=position,column=5).value
                    balance=sheet.cell(row=position,column=6).value
                    status=sheet.cell(row=position,column=7).value
                    print("Valid user")
                    break
                else:
                    chance-=1
        else:
            print("Sorry you cannot perform any transactions your account status is blocked")
        if chance==0:
            sheet.cell(position,7).value='Blocked'
            print("Dear customer sorry but your account has been blocked\nQuiting the program!!")
            wb.save('SITIP19_KIIT_Python_FinalProject1_Workbook.xlsx')
        else:
            otp=random.randint(1000,9999)
            print("Please wait..")
            fromaddr=""                         #enter your email address inside quotes
            toaddr=mailID
            msg=MIMEMultipart()
            msg['From']=fromaddr
            msg['To']=toaddr
            msg['Subject']='Account Verification'
            body="Sir,this is your OTP Number: {}".format(otp)
            msg.attach(MIMEText(body,'plain'))
            server=smtplib.SMTP('smtp.gmail.com',port=587)
            server.starttls()
            server.login(fromaddr,'')                       #enter your email's password inside quotes
            text=msg.as_string()
            server.sendmail(fromaddr,toaddr,text)
            server.quit()
            time.sleep(5)
            print("An OTP has been sent to your mail. Plese Check!")
            while chance1!=0:
                otpv=int(input("Sir please enter the OTP that you received via mail: "))
                if otpv==otp:
                    print("Valid OTP User verified")
                    break
                else:
                    print("Invalid OTP")
                    chance1-=1
            if chance1==0:
                sheet.cell(position,7).value='Blocked'
                print("Dear customer sorry but your account has been blocked\nQuiting the program!!")
                wb.save('SITIP19_KIIT_Python_FinalProject1_Workbook.xlsx')
            else:
                print("Choose Transaction type:")
                while chance2!=0:
                    transtype=input("C-->Credit money\nD-->Debit money: ")
                    if transtype=='c' or transtype=='d' or transtype=='C' or transtype=='D':
                        print("Your name: %s\nCurrent Balance: %.2f /-"%(fullname,float(balance)))
                        if transtype=='c' or transtype=='C':
                            amtc=int(input("Enter amount to be credited: "))
                            print("Successful Transaction of Rs. %.2f /-\nCurrent Balance: %.2f /-"%(float(amtc),(float(balance)+float(amtc))))
                            sheet.cell(position,6).value=float(balance)+float(amtc)
                            amtd=0
                            wb.save('SITIP19_KIIT_Python_FinalProject1_Workbook.xlsx')
                        else:
                            amtd=int(input("Enter amount to be debited: "))
                            if int(amtd)<int(balance):
                                print("Successful Transaction of Rs. %.2f /-\nCurrent Balance: %.2f /-"%(float(amtd),(float(balance)-float(amtd))))
                                sheet.cell(position,6).value=float(balance)-float(amtd)
                                amtc=0
                                wb.save('SITIP19_KIIT_Python_FinalProject1_Workbook.xlsx')
                            else:
                                print("Sorry! Not enough balance to fulfil your transaction")
                                amtc=amtd=0
                        print("Your Account Details:\nName: %s"%(fullname))
                        print("Transaction Date and Time: ",now.strftime("%Y-%m-%d %H:%M:%S"))
                        print("Amount Withdrawn: Rs. %.2f"%(amtd if amtd else 0))
                        print("Amount Deposited: Rs. %.2f"%(amtc if amtc else 0))
                        print("Balance: Rs. %.2f"%(float(sheet.cell(position,6).value)))
                        break
                    else:
                        print("Invalid input")
                        chance2-=1
                if chance2==0:
                    sheet.cell(position,7).value='Blocked'
                    print("Dear customer Sorry but your account has been blocked\nQuiting the program!!")
                    wb.save('SITIP19_KIIT_Python_FinalProject1_Workbook.xlsx')
except ValueError:
    print("Invalid input entered")
